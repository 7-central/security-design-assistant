"""Comprehensive end-to-end integration test for the pipeline."""
import asyncio
import json
import time
from pathlib import Path
from unittest.mock import Mock, patch

import pytest
from openpyxl import load_workbook

from src.agents.excel_generation_agent import ExcelGenerationAgent
from src.agents.schedule_agent_v2 import ScheduleAgentV2
from src.models.job import Job, JobStatus
from src.storage.local_storage import LocalStorage


@pytest.fixture
def test_drawing_path():
    """Get path to test drawing."""
    return Path("tests/fixtures/drawings/variations/01_different_text_sizes.pdf")


@pytest.fixture
def mock_storage(tmp_path):
    """Create a mock storage with temp directory."""
    with patch.dict('os.environ', {'LOCAL_OUTPUT_DIR': str(tmp_path)}):
        storage = LocalStorage()
        return storage


@pytest.fixture
def test_job():
    """Create a test job instance."""
    now = int(time.time())
    return Job(
        job_id=f"test_job_{now}",
        client_name="TestClient",
        project_name="IntegrationTest",
        status=JobStatus.QUEUED,
        created_at=now,
        updated_at=now,
        metadata={
            "file_name": "test_drawing.pdf",
            "file_size_mb": 1.5
        }
    )


async def upload_and_process_drawing(storage, job, drawing_path):
    """Upload drawing and initiate processing.

    Args:
        storage: Storage interface
        job: Job instance
        drawing_path: Path to drawing file

    Returns:
        Uploaded file path
    """
    # Read drawing file
    with open(drawing_path, 'rb') as f:
        drawing_content = f.read()

    # Save to storage
    file_key = f"{job.client_name}/{job.project_name}/{job.job_id}/drawing.pdf"
    uploaded_path = await storage.save_file(file_key, drawing_content)

    # Update job with file path
    job.file_path = uploaded_path
    job.status = JobStatus.PROCESSING
    await storage.save_job_status(job.job_id, job.to_dict())

    return uploaded_path


async def wait_for_job_completion(storage, job, timeout=60):
    """Wait for job to complete with timeout.

    Args:
        storage: Storage interface
        job: Job instance
        timeout: Maximum wait time in seconds

    Returns:
        Final job status

    Raises:
        TimeoutError: If job doesn't complete within timeout
    """
    start_time = time.time()

    while time.time() - start_time < timeout:
        job_status = await storage.get_job_status(job.job_id)

        if job_status and job_status.get('status') in [JobStatus.COMPLETED, JobStatus.FAILED]:
            return job_status

        await asyncio.sleep(1)

    raise TimeoutError(f"Job {job.job_id} did not complete within {timeout} seconds")


def validate_job_response(job_response):
    """Validate all required fields in job response.

    Args:
        job_response: Job status dictionary

    Returns:
        List of validation errors (empty if valid)
    """
    errors = []

    # Check required top-level fields
    required_fields = ['job_id', 'status', 'client_name', 'project_name', 'created_at']
    for field in required_fields:
        if field not in job_response:
            errors.append(f"Missing required field: {field}")

    # Check processing results
    if 'processing_results' not in job_response:
        errors.append("Missing processing_results")
    else:
        results = job_response['processing_results']

        # Check schedule agent results
        if 'schedule_agent' in results:
            schedule = results['schedule_agent']
            if 'components' not in schedule:
                errors.append("Missing components in schedule_agent results")
            elif not isinstance(schedule['components'], list):
                errors.append("Components should be a list")

        # Check excel generation results
        if 'excel_generation' in results:
            excel = results['excel_generation']
            if 'excel_path' not in excel:
                errors.append("Missing excel_path in excel_generation results")

    return errors


async def verify_excel_generation(storage, job_response):
    """Verify Excel file was generated and is valid.

    Args:
        storage: Storage interface
        job_response: Job status dictionary

    Returns:
        Tuple of (is_valid, error_message)
    """
    try:
        # Get Excel path from results
        excel_path = job_response.get('processing_results', {}).get('excel_generation', {}).get('excel_path')

        if not excel_path:
            return False, "No Excel path in job response"

        # Check file exists
        if not await storage.file_exists(excel_path):
            return False, f"Excel file does not exist: {excel_path}"

        # Try to load the Excel file
        excel_content = await storage.get_file(excel_path)

        # Save to temp file and validate
        temp_excel = Path("/tmp/test_excel.xlsx")
        temp_excel.write_bytes(excel_content)

        wb = load_workbook(temp_excel)

        # Check expected sheets
        expected_sheets = ['Schedule', 'Summary']
        for sheet in expected_sheets:
            if sheet not in wb.sheetnames:
                return False, f"Missing expected sheet: {sheet}"

        # Check Schedule sheet has data
        schedule_sheet = wb['Schedule']
        if schedule_sheet.max_row < 2:  # Header + at least one data row
            return False, "Schedule sheet has no data rows"

        return True, None

    except Exception as e:
        return False, f"Error validating Excel: {e!s}"


def data_integrity_checks(job_response):
    """Check data integrity and consistency.

    Args:
        job_response: Job status dictionary

    Returns:
        List of integrity issues (empty if good)
    """
    issues = []

    try:
        results = job_response.get('processing_results', {})

        # Check component data integrity
        if 'schedule_agent' in results:
            components = results['schedule_agent'].get('components', [])

            for i, component in enumerate(components):
                # Check required component fields
                required = ['component_id', 'type', 'location']
                for field in required:
                    if field not in component:
                        issues.append(f"Component {i} missing field: {field}")

                # Check component ID format (should start with A-)
                comp_id = component.get('component_id', '')
                if not comp_id.startswith('A-'):
                    issues.append(f"Component {comp_id} has invalid prefix (should be A-)")

                # Check location is dict with x, y
                location = component.get('location', {})
                if not isinstance(location, dict):
                    issues.append(f"Component {comp_id} location is not a dict")
                elif 'x' not in location or 'y' not in location:
                    issues.append(f"Component {comp_id} location missing x or y")

        # Check timestamps are valid
        created_at = job_response.get('created_at')
        if created_at:
            if not isinstance(created_at, (int, float)):
                issues.append(f"Invalid created_at timestamp: {created_at}")
            elif created_at > time.time():
                issues.append(f"created_at is in the future: {created_at}")

    except Exception as e:
        issues.append(f"Exception during integrity check: {e!s}")

    return issues


@pytest.mark.integration
@pytest.mark.asyncio
class TestPipelineEndToEnd:
    """Comprehensive end-to-end integration tests."""

    async def test_upload_and_process_drawing(self, mock_storage, test_job, test_drawing_path):
        """Test: Upload drawing → Wait for completion → Validate output."""
        # Upload drawing
        uploaded_path = await upload_and_process_drawing(
            mock_storage, test_job, test_drawing_path
        )

        assert uploaded_path is not None
        assert await mock_storage.file_exists(uploaded_path)

        # Mock the pipeline processing
        with patch('src.agents.base_agent_v2.genai.Client') as mock_client:
            # Mock Gemini responses
            mock_response = Mock()
            mock_response.text = json.dumps({
                "components": [
                    {
                        "component_id": "A-200",
                        "type": "door",
                        "location": {"x": 100, "y": 200},
                        "description": "Door with 8pt text",
                        "has_reader": True,
                        "has_rex": True
                    },
                    {
                        "component_id": "A-201",
                        "type": "door",
                        "location": {"x": 200, "y": 200},
                        "description": "Door with 10pt text",
                        "has_reader": True,
                        "has_rex": True
                    }
                ]
            })
            mock_client.return_value.models.generate_content.return_value = mock_response

            # Process through pipeline
            schedule_agent = ScheduleAgentV2(mock_storage, test_job)
            schedule_agent._client = mock_client.return_value

            schedule_result = await schedule_agent.process({
                "drawing_path": uploaded_path,
                "pages": [{"page_num": 1, "content": "test"}]
            })

            # Update job with results
            test_job.processing_results = {
                "schedule_agent": schedule_result
            }

            # Generate Excel
            excel_agent = ExcelGenerationAgent(mock_storage, test_job)
            excel_agent._client = mock_client.return_value

            # Mock Excel generation response
            excel_response = Mock()
            excel_response.text = "Excel content generated"
            excel_response.files = [Mock(name="schedule.xlsx", content=b"fake excel content")]
            mock_client.return_value.models.generate_content.return_value = excel_response

            excel_result = await excel_agent.process({
                "components": schedule_result["components"]
            })

            test_job.processing_results["excel_generation"] = excel_result

            # Mark job as completed
            test_job.status = JobStatus.COMPLETED
            await mock_storage.save_job_status(test_job.job_id, test_job.to_dict())

        # Wait for completion
        final_status = await wait_for_job_completion(mock_storage, test_job, timeout=10)

        assert final_status['status'] == JobStatus.COMPLETED

        # Validate response
        validation_errors = validate_job_response(final_status)
        assert len(validation_errors) == 0, f"Validation errors: {validation_errors}"

    async def test_verify_all_expected_fields(self, mock_storage, test_job):
        """Test: Verify all expected fields in response."""
        # Create a complete job response
        job_response = {
            "job_id": test_job.job_id,
            "status": JobStatus.COMPLETED,
            "client_name": test_job.client_name,
            "project_name": test_job.project_name,
            "created_at": test_job.created_at,
            "updated_at": int(time.time()),
            "metadata": {
                "file_name": "test.pdf",
                "total_pages": 1
            },
            "processing_results": {
                "schedule_agent": {
                    "completed": True,
                    "components": [
                        {
                            "component_id": "A-001",
                            "type": "door",
                            "location": {"x": 100, "y": 200}
                        }
                    ]
                },
                "excel_generation": {
                    "completed": True,
                    "excel_path": "test/excel.xlsx"
                }
            }
        }

        # Validate all fields
        errors = validate_job_response(job_response)
        assert len(errors) == 0

    async def test_excel_file_generation(self, mock_storage, test_job, tmp_path):
        """Test: Check Excel file is generated."""
        # Create mock Excel file
        excel_path = f"{test_job.client_name}/{test_job.project_name}/excel.xlsx"

        # Create a simple Excel file for testing
        from openpyxl import Workbook
        wb = Workbook()

        # Create Schedule sheet
        schedule = wb.active
        schedule.title = "Schedule"
        schedule.append(["Component ID", "Type", "Location"])
        schedule.append(["A-001", "Door", "Room 101"])

        # Create Summary sheet
        summary = wb.create_sheet("Summary")
        summary.append(["Total Components", "1"])

        # Save to temp location
        temp_excel = tmp_path / "test.xlsx"
        wb.save(temp_excel)

        # Save to storage
        with open(temp_excel, 'rb') as f:
            await mock_storage.save_file(excel_path, f.read())

        # Create job response
        job_response = {
            "processing_results": {
                "excel_generation": {
                    "excel_path": excel_path
                }
            }
        }

        # Verify Excel
        is_valid, error = await verify_excel_generation(mock_storage, job_response)
        assert is_valid, f"Excel validation failed: {error}"

    async def test_data_integrity(self, test_job):
        """Test: Confirm no data corruption."""
        # Test valid data
        valid_response = {
            "job_id": test_job.job_id,
            "created_at": int(time.time()),
            "processing_results": {
                "schedule_agent": {
                    "components": [
                        {
                            "component_id": "A-001",
                            "type": "door",
                            "location": {"x": 100, "y": 200}
                        },
                        {
                            "component_id": "A-002",
                            "type": "reader",
                            "location": {"x": 150, "y": 200}
                        }
                    ]
                }
            }
        }

        issues = data_integrity_checks(valid_response)
        assert len(issues) == 0

        # Test corrupted data
        corrupted_response = {
            "job_id": test_job.job_id,
            "created_at": int(time.time()) + 10000,  # Future timestamp
            "processing_results": {
                "schedule_agent": {
                    "components": [
                        {
                            "component_id": "C-001",  # Wrong prefix
                            "type": "door",
                            "location": "wrong format"  # Should be dict
                        },
                        {
                            "component_id": "A-002",
                            # Missing required fields
                            "location": {"x": 150}  # Missing y
                        }
                    ]
                }
            }
        }

        issues = data_integrity_checks(corrupted_response)
        assert len(issues) > 0
        assert any("prefix" in issue for issue in issues)
        assert any("location" in issue for issue in issues)
        assert any("future" in issue for issue in issues)
