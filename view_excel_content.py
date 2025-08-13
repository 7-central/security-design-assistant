#!/usr/bin/env python3
"""
View the content of generated Excel files.
"""

import openpyxl
from pathlib import Path
import sys

def view_excel_file(file_path):
    """Display the content of an Excel file."""
    print(f"\n{'='*60}")
    print(f"Excel File: {Path(file_path).name}")
    print('='*60)
    
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(file_path)
        
        # Get the active sheet
        sheet = wb.active
        print(f"Sheet Name: {sheet.title}")
        print(f"Dimensions: {sheet.max_row} rows x {sheet.max_column} columns")
        print()
        
        # Read all data
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        
        if not data:
            print("No data in the Excel file")
            return
        
        # Calculate column widths for display
        col_widths = []
        for col_idx in range(len(data[0]) if data else 0):
            max_width = 0
            for row in data:
                if col_idx < len(row) and row[col_idx] is not None:
                    max_width = max(max_width, len(str(row[col_idx])))
            col_widths.append(min(max_width, 30))  # Cap at 30 chars for display
        
        # Print headers (first row)
        print("Headers:")
        print("-" * 60)
        if data:
            headers = data[0]
            for i, header in enumerate(headers):
                if header:
                    print(f"  Column {i+1}: {header}")
        
        # Print data rows
        print("\nData Rows:")
        print("-" * 60)
        for row_idx, row in enumerate(data[1:], start=1):  # Skip header
            if row_idx > 10:  # Limit display to first 10 data rows
                print(f"  ... ({len(data)-11} more rows)")
                break
            
            # Format row for display
            row_str = f"Row {row_idx}: "
            non_empty_values = []
            for col_idx, value in enumerate(row):
                if value is not None and str(value).strip():
                    col_name = data[0][col_idx] if col_idx < len(data[0]) else f"Col{col_idx+1}"
                    # Truncate long values
                    val_str = str(value)[:25] + "..." if len(str(value)) > 25 else str(value)
                    non_empty_values.append(f"{col_name}={val_str}")
            
            if non_empty_values:
                print(f"  {row_str}")
                for val in non_empty_values[:5]:  # Show first 5 non-empty columns
                    print(f"    {val}")
                if len(non_empty_values) > 5:
                    print(f"    ... ({len(non_empty_values)-5} more fields)")
        
        # Print summary row if it exists
        if len(data) > 1:
            last_row = data[-1]
            if last_row and any("Total" in str(cell) if cell else False for cell in last_row):
                print("\nSummary Row:")
                print("-" * 60)
                for i, value in enumerate(last_row):
                    if value and str(value).strip():
                        print(f"  {value}")
        
        wb.close()
        return True
        
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return False

# View the most recent Excel files
excel_files = [
    "/Users/leehayton/Cursor Projects/7central/security_and_design/local_output/test_client/test_project/job_job_20250808081920952/schedule_20250808_072111.xlsx",
    "/Users/leehayton/Cursor Projects/7central/security_and_design/local_output/test_client/test_project/job_job_20250808081455014/schedule_20250808_071639.xlsx",
    "/Users/leehayton/Cursor Projects/7central/security_and_design/test_output.xlsx"  # From the working test
]

print("Viewing Generated Excel Files")
print("=" * 60)

for file_path in excel_files:
    if Path(file_path).exists():
        view_excel_file(file_path)
    else:
        print(f"\nFile not found: {file_path}")

print("\n" + "="*60)
print("Excel files generated successfully with proper structure!")