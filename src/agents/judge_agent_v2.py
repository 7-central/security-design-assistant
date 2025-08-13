"""Judge Agent V2 for evaluating extraction quality using Google GenAI SDK."""
import json
from pathlib import Path
from typing import Any

import openpyxl
from google import genai

from src.agents.base_agent_v2 import BaseAgentV2


class JudgeAgentV2(BaseAgentV2):
    """Agent for evaluating the quality of component extraction and Excel generation."""

    def __init__(self, storage, job):
        """Initialize Judge Agent."""
        super().__init__(storage, job)
        self.model_name = "models/gemini-2.5-pro"  # Using Gemini 2.5 Pro for evaluation
        self.prompt_file = Path("src/config/prompts/judge_prompt.txt")

    def _load_prompt_template(self) -> str:
        """Load the judge prompt template from file."""
        try:
            if self.prompt_file.exists():
                return self.prompt_file.read_text()
            else:
                # Fallback to inline prompt if file doesn't exist yet
                self.log_structured("warning", "Judge prompt file not found, using inline prompt")
                return self._get_fallback_prompt()
        except Exception as e:
            self.log_structured("error", f"Error loading prompt template: {e}")
            return self._get_fallback_prompt()

    def _get_fallback_prompt(self) -> str:
        """Get fallback prompt if template file is not available."""
        return """
You are evaluating the quality of a security drawing processing pipeline that extracts access control components.

Pipeline Scope: This system extracts access control components (readers, exit buttons, door controllers, etc.)
from security drawings and generates Excel schedules.

{drawing_info}

{context_info}

{components_info}

{excel_info}

Evaluation Framework:
Please evaluate the extraction quality using these 5 consistent questions:

1. **Completeness**: Looking at the drawing, are there obvious access control components that were missed?
2. **Correctness**: Are the extracted components correctly identified and classified?
3. **Context Usage**: Did the system appropriately use the provided context to enhance extraction?
4. **Spatial Understanding**: Are components correctly associated (e.g., readers with doors)?
5. **False Positives**: Are there any components in the schedule that don't appear in the drawing?

Provide your evaluation in the following JSON format:
{{
    "overall_assessment": "[Good/Fair/Poor] performance with clear reasoning",
    "completeness": "Description of what was found vs missed",
    "correctness": "Assessment of accuracy in identification and classification",
    "context_usage": "How well context was applied",
    "spatial_understanding": "Quality of spatial relationships",
    "false_positives": "Any incorrectly identified components",
    "improvement_suggestions": [
        "Specific actionable suggestion 1",
        "Specific actionable suggestion 2"
    ]
}}"""

    def _build_evaluation_prompt(
        self, drawing_path: Path | None, components: list[dict], excel_path: Path | None, context: dict | None
    ) -> tuple[str, list[genai.types.File]]:
        """Build prompt for quality evaluation with file uploads.

        Returns:
            Tuple of (prompt_text, list_of_files_to_upload)
        """
        files_to_upload = []

        # Upload drawing if available
        drawing_info = "Drawing: Not provided"
        if drawing_path and drawing_path.exists():
            try:
                drawing_file = self.client.files.upload(path=str(drawing_path))
                files_to_upload.append(drawing_file)
                drawing_info = f"Drawing: Uploaded {drawing_path.name} for analysis"
            except Exception as e:
                self.log_structured("warning", f"Could not upload drawing: {e}")
                drawing_info = f"Drawing: Failed to upload ({e})"

        # Format context information
        context_info = "Context: No context provided"
        if context:
            context_sections = []
            for key, value in context.items():
                if value:
                    context_sections.append(f"  - {key}: {value}")
            if context_sections:
                context_info = "Context used in extraction:\n" + "\n".join(context_sections)

        # Format components information
        components_info = f"Extracted Components: {len(components)} total\n"
        if components:
            # Show first 5 components as examples
            components_info += "Sample components:\n"
            for comp in components[:5]:
                components_info += (
                    f"  - ID: {comp.get('id', 'N/A')}, "
                    f"Type: {comp.get('type', 'N/A')}, "
                    f"Location: {comp.get('location', 'N/A')}\
"
                )
            if len(components) > 5:
                components_info += f"  ... and {len(components) - 5} more components\n"

            # Add component statistics
            type_counts = {}
            for comp in components:
                comp_type = comp.get("type", "unknown")
                type_counts[comp_type] = type_counts.get(comp_type, 0) + 1
            components_info += "\nComponent type distribution:\n"
            for comp_type, count in sorted(type_counts.items()):
                components_info += f"  - {comp_type}: {count}\n"

        # Format Excel information
        excel_info = "Excel File: Not generated"
        if excel_path and excel_path.exists():
            try:
                # Read Excel to get sheet information
                wb = openpyxl.load_workbook(excel_path, read_only=True)
                sheet_info = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    row_count = ws.max_row
                    col_count = ws.max_column
                    sheet_info.append(f"  - {sheet_name}: {row_count} rows x {col_count} columns")
                wb.close()
                excel_info = "Excel File: Generated successfully\nSheets:\n" + "\n".join(sheet_info)
            except Exception as e:
                excel_info = f"Excel File: Generated but could not read ({e})"

        # Load and format prompt template
        prompt_template = self._load_prompt_template()
        prompt = prompt_template.format(
            drawing_info=drawing_info, context_info=context_info, components_info=components_info, excel_info=excel_info
        )

        return prompt, files_to_upload

    def _validate_evaluation(self, evaluation: dict) -> dict:
        """Validate and ensure all required fields are present in evaluation."""
        required_fields = [
            "overall_assessment",
            "completeness",
            "correctness",
            "context_usage",
            "spatial_understanding",
            "false_positives",
            "improvement_suggestions",
        ]

        # Add missing fields with defaults
        for field in required_fields:
            if field not in evaluation:
                if field == "improvement_suggestions":
                    evaluation[field] = ["Unable to generate specific suggestions"]
                elif field == "overall_assessment":
                    evaluation[field] = "Fair performance - evaluation incomplete"
                else:
                    evaluation[field] = "Not evaluated"

        # Ensure improvement_suggestions is a list
        if not isinstance(evaluation.get("improvement_suggestions"), list):
            evaluation["improvement_suggestions"] = [str(evaluation.get("improvement_suggestions", "No suggestions"))]

        return evaluation

    def _parse_evaluation_response(self, response: str) -> dict[str, Any]:
        """Parse evaluation response with error handling."""
        try:
            # Extract JSON from response
            if "```json" in response:
                json_str = response.split("```json")[1].split("```")[0].strip()
            elif "{" in response and "}" in response:
                start = response.find("{")
                end = response.rfind("}") + 1
                json_str = response[start:end]
            else:
                json_str = response

            evaluation = json.loads(json_str)

            # Validate and fill missing fields
            evaluation = self._validate_evaluation(evaluation)

            return evaluation

        except json.JSONDecodeError as e:
            self.log_structured("warning", f"Failed to parse evaluation response: {e}")
            # Return a basic evaluation with required fields
            return {
                "overall_assessment": "Poor performance - evaluation failed",
                "completeness": "Unable to evaluate",
                "correctness": "Unable to evaluate",
                "context_usage": "Unable to evaluate",
                "spatial_understanding": "Unable to evaluate",
                "false_positives": "Unable to evaluate",
                "improvement_suggestions": ["Fix evaluation errors and retry"],
                "error": str(e),
            }

    async def evaluate_extraction(
        self, drawing_path: Path | None, context: dict | None, components: list[dict], excel_path: Path | None
    ) -> dict[str, Any]:
        """Evaluate the quality of extraction and generation.

        Args:
            drawing_path: Path to original drawing PDF
            context: Context data used in extraction
            components: List of extracted components
            excel_path: Path to generated Excel file

        Returns:
            Dictionary with evaluation results
        """
        self.log_structured("info", "Starting judge evaluation")

        try:
            # Build evaluation prompt with file uploads
            prompt, files = self._build_evaluation_prompt(drawing_path, components, excel_path, context)

            # Generate evaluation using Gemini
            if files:
                response = await self._generate_with_files(prompt, files)
            else:
                response = await self._generate_with_retry(prompt)

            # Parse and validate evaluation
            evaluation = self._parse_evaluation_response(response)

            # Log evaluation results
            self.log_structured(
                "info", f"Evaluation complete - Assessment: {evaluation.get('overall_assessment', 'Unknown')}"
            )

            # Log assessment trend for analytics
            assessment = evaluation.get("overall_assessment", "")
            if "Good" in assessment:
                self.log_structured("info", "Assessment trend: Good")
            elif "Fair" in assessment:
                self.log_structured("info", "Assessment trend: Fair")
            elif "Poor" in assessment:
                self.log_structured("info", "Assessment trend: Poor")

            # Log common improvement suggestions
            suggestions = evaluation.get("improvement_suggestions", [])
            if suggestions:
                self.log_structured("info", f"Improvement suggestions: {', '.join(suggestions[:2])}")

            return evaluation

        except Exception as e:
            self.log_structured("error", f"Judge evaluation failed: {e}")
            return self._validate_evaluation(
                {"overall_assessment": f"Poor performance - evaluation error: {e}", "error": str(e)}
            )

    async def _generate_with_files(self, prompt: str, files: list[genai.types.File]) -> str:
        """Generate response with uploaded files."""
        try:
            # Convert File objects to proper Part format (like Schedule Agent does)
            from google.genai import types

            file_parts = []
            for file in files:
                file_part = types.Part(
                    file_data=types.FileData(file_uri=file.uri, mime_type=file.mime_type or "application/pdf")
                )
                file_parts.append(file_part)

            response = self.client.models.generate_content(
                model=self.model_name,
                contents=[
                    *file_parts,  # Include uploaded files as Parts
                    prompt,
                ],
                config=genai.types.GenerateContentConfig(temperature=0.3, top_p=0.9, max_output_tokens=4096),
            )
            # Ensure we return text, not the response object
            if hasattr(response, "text"):
                return response.text
            else:
                return str(response)
        except Exception as e:
            self.log_structured("error", f"Generation with files failed: {e}")
            raise

    async def process(self, input_data: dict[str, Any]) -> dict[str, Any]:
        """Process evaluation request from pipeline.

        Args:
            input_data: Dictionary with paths to artifacts

        Returns:
            Dictionary with evaluation results and next stage
        """
        self.log_structured("info", "Processing judge evaluation request")

        try:
            # Load artifacts from storage
            drawing_path = None
            excel_path = None
            components = []
            context = None

            # Get drawing path if available
            if "drawing_file" in input_data and input_data["drawing_file"]:
                drawing_path = Path(input_data["drawing_file"])

            # Get Excel path if available
            if "excel_file" in input_data and input_data["excel_file"]:
                excel_path = Path(input_data["excel_file"])

            # Get components
            components = self._extract_components(input_data)

            # Get context if available
            context = input_data.get("context")

            # Run evaluation
            evaluation = await self.evaluate_extraction(drawing_path, context, components, excel_path)

            # Save evaluation checkpoint
            await self.save_checkpoint("evaluation", {"evaluation": evaluation, "timestamp": self._get_timestamp()})

            # Include evaluation in job metadata
            result = {
                "evaluation": evaluation,
                "next_stage": "complete",
                "metadata": {
                    "overall_assessment": evaluation.get("overall_assessment"),
                    "suggestions_count": len(evaluation.get("improvement_suggestions", [])),
                },
            }

            return result

        except Exception as e:
            self.log_structured("error", f"Judge process failed: {e}")
            return {
                "evaluation": self._validate_evaluation(
                    {"overall_assessment": f"Poor performance - process error: {e}", "error": str(e)}
                ),
                "next_stage": "complete",
            }

    def _get_timestamp(self) -> str:
        """Get current timestamp in ISO format."""
        from datetime import datetime

        return datetime.utcnow().isoformat() + "Z"

    def _extract_components(self, input_data: dict[str, Any]) -> list[dict]:
        """Extract components from various input formats."""
        # Try different locations where components might be stored
        if "components" in input_data:
            return input_data["components"]
        elif "schedule_data" in input_data and "components" in input_data["schedule_data"]:
            return input_data["schedule_data"]["components"]
        elif "pages" in input_data:
            components = []
            for page in input_data["pages"]:
                if "components" in page:
                    components.extend(page["components"])
            return components
        else:
            return []
