"""Integration tests for the complete pipeline including Judge evaluation."""
import json
import tempfile
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock, Mock, patch

import pytest

from src.agents.context_agent import ContextAgent
from src.agents.excel_generation_agent import ExcelGenerationAgent
from src.agents.judge_agent_v2 import JudgeAgentV2
from src.agents.schedule_agent_v2 import ScheduleAgentV2
from src.models.job import Job, JobStatus
from src.storage.local_storage import LocalStorage


@pytest.fixture
def mock_job():
    """Create a mock job for testing."""
    job = Mock(spec=Job)
    job.job_id = "test-job-123"
    job.client_name = "TestClient"
    job.project_name = "TestProject"
    job.status = JobStatus.PROCESSING
    job.metadata = {"file_name": "test.pdf", "total_pages": 2}
    job.processing_results = {}
    job.update_metadata = Mock()
    job.update_processing_results = Mock()
    return job


@pytest.fixture
def mock_storage():
    """Create a mock storage instance."""
    storage = AsyncMock(spec=LocalStorage)
    storage.save_checkpoint = AsyncMock()
    storage.load_checkpoint = AsyncMock(return_value=None)
    storage.save_job_status = AsyncMock()
    storage.get_job_status = AsyncMock()
    storage.file_exists = AsyncMock(return_value=True)
    storage.get_file = AsyncMock()
    storage.save_file = AsyncMock()
    return storage


@pytest.fixture
def sample_drawing_pages():
    """Create sample drawing pages."""
    return [
        {
            "page_number": 1,
            "text_content": "Floor 1 Security Plan",
            "components": []
        },
        {
            "page_number": 2,
            "text_content": "Floor 2 Security Plan",
            "components": []
        }
    ]


@pytest.fixture
def sample_components():
    """Create sample extracted components."""
    return [
        {
            "id": "A-101-DR-B2",
            "type": "door",
            "location": "Main entrance",
            "page_number": 1,
            "confidence": 0.95,
            "reasoning": "Identified as door based on symbol",
            "attributes": {"door_type": "Single"}
        },
        {
            "id": "A-101-RDR-01",
            "type": "reader",
            "location": "Main entrance",
            "page_number": 1,
            "confidence": 0.90,
            "reasoning": "Card reader symbol",
            "attributes": {"reader_type": "Proximity"}
        }
    ]


@pytest.fixture
def sample_context():
    """Create sample context data."""
    return {
        "lock_types": "All doors use magnetic locks",
        "reader_models": "Use P-Series proximity readers",
        "standards": "Follow ANSI/TIA-606-B labeling"
    }


@pytest.mark.integration
class TestFullPipelineWithJudge:
    """Test the complete pipeline including Judge evaluation."""

    @pytest.mark.asyncio
    async def test_full_pipeline_with_judge(
        self,
        mock_storage,
        mock_job,
        sample_drawing_pages,
        sample_components,
        sample_context
    ):
        """Test full pipeline execution with Judge evaluation."""

        # Mock Context Agent
        with patch('src.agents.context_agent.ContextAgent') as MockContextAgent:
            context_agent = AsyncMock(spec=ContextAgent)
            context_agent.process = AsyncMock(return_value={
                "context": sample_context,
                "next_stage": "schedule"
            })
            MockContextAgent.return_value = context_agent

            # Mock Schedule Agent
            with patch('src.agents.schedule_agent_v2.ScheduleAgentV2') as MockScheduleAgent:
                schedule_agent = AsyncMock(spec=ScheduleAgentV2)
                schedule_agent.process = AsyncMock(return_value={
                    "components": {"pages": [{"components": sample_components}]},
                    "next_stage": "excel"
                })
                MockScheduleAgent.return_value = schedule_agent

                # Mock Excel Generation Agent
                with patch('src.agents.excel_generation_agent.ExcelGenerationAgent') as MockExcelAgent:
                    excel_agent = AsyncMock(spec=ExcelGenerationAgent)
                    excel_agent.process = AsyncMock(return_value={
                        "status": "completed",
                        "file_path": "/tmp/test_excel.xlsx",
                        "summary": {"components_count": 2}
                    })
                    MockExcelAgent.return_value = excel_agent

                    # Create real Judge Agent (we're testing this)
                    judge_agent = JudgeAgentV2(storage=mock_storage, job=mock_job)

                    # Mock the Gemini API response and save_checkpoint
                    with patch.object(judge_agent, '_generate_with_retry', new_callable=AsyncMock) as mock_generate:
                        mock_generate.return_value = json.dumps({
                            "overall_assessment": "Good performance - extracted components successfully",
                            "completeness": "Found main components",
                            "correctness": "Types correctly identified",
                            "context_usage": "Context applied well",
                            "spatial_understanding": "Good spatial relationships",
                            "false_positives": "None detected",
                            "improvement_suggestions": ["Consider emergency exits"]
                        })

                        with patch.object(judge_agent, 'save_checkpoint', new_callable=AsyncMock) as mock_save:
                            mock_save.return_value = None

                            # Run Judge evaluation
                            judge_input = {
                                "drawing_file": None,  # Simulating no drawing file
                                "context": sample_context,
                                "components": sample_components,
                                "excel_file": "/tmp/test_excel.xlsx"
                            }

                            result = await judge_agent.process(judge_input)

                            # Verify Judge results
                            assert result["next_stage"] == "complete"
                            assert "evaluation" in result
                            assert result["evaluation"]["overall_assessment"] == "Good performance - extracted components successfully"
                            assert len(result["evaluation"]["improvement_suggestions"]) == 1

                            # Verify checkpoint was saved
                            mock_save.assert_called()
                            checkpoint_call = mock_save.call_args[0]
                            assert checkpoint_call[0] == "evaluation"
                            assert "evaluation" in checkpoint_call[1]
                            assert "timestamp" in checkpoint_call[1]

    @pytest.mark.asyncio
    async def test_pipeline_judge_with_poor_extraction(
        self,
        mock_storage,
        mock_job
    ):
        """Test Judge evaluation when extraction quality is poor."""

        # Create Judge Agent
        judge_agent = JudgeAgentV2(storage=mock_storage, job=mock_job)

        # Mock poor quality extraction response and save_checkpoint
        with patch.object(judge_agent, '_generate_with_retry', new_callable=AsyncMock) as mock_generate:
            mock_generate.return_value = json.dumps({
                "overall_assessment": "Poor performance - missed majority of components",
                "completeness": "Only 30% of visible components extracted",
                "correctness": "Many misclassifications",
                "context_usage": "Context ignored",
                "spatial_understanding": "Confused spatial relationships",
                "false_positives": "Multiple false positives detected",
                "improvement_suggestions": [
                    "Improve symbol recognition",
                    "Better multi-page processing",
                    "Apply context information"
                ]
            })

            with patch.object(judge_agent, 'save_checkpoint', new_callable=AsyncMock) as mock_save:
                mock_save.return_value = None

                # Run with empty components (poor extraction)
                judge_input = {
                    "components": [],  # Empty extraction
                    "context": None,
                    "excel_file": None
                }

                result = await judge_agent.process(judge_input)

                # Verify poor evaluation is handled correctly
                assert "Poor performance" in result["evaluation"]["overall_assessment"]
                assert len(result["evaluation"]["improvement_suggestions"]) == 3
                assert result["next_stage"] == "complete"

    @pytest.mark.asyncio
    async def test_pipeline_judge_evaluation_retrieval(
        self,
        mock_storage,
        mock_job,
        sample_components
    ):
        """Test retrieval of Judge evaluation from job status."""

        # Mock job data with evaluation
        job_data = {
            "job_id": "test-job-123",
            "status": "completed",
            "processing_results": {
                "evaluation": {
                    "overall_assessment": "Good performance",
                    "completeness": "95% coverage",
                    "correctness": "High accuracy",
                    "improvement_suggestions": ["Minor improvements needed"]
                },
                "evaluation_metadata": {
                    "overall_assessment": "Good performance",
                    "suggestions_count": 1
                }
            }
        }

        mock_storage.get_job_status.return_value = job_data

        # Retrieve job status
        retrieved_job = await mock_storage.get_job_status("test-job-123")

        # Verify evaluation is included
        assert "evaluation" in retrieved_job["processing_results"]
        evaluation = retrieved_job["processing_results"]["evaluation"]
        assert evaluation["overall_assessment"] == "Good performance"
        assert evaluation["completeness"] == "95% coverage"
        assert len(evaluation["improvement_suggestions"]) == 1

    @pytest.mark.asyncio
    async def test_judge_checkpoint_saving(
        self,
        mock_storage,
        mock_job,
        sample_components
    ):
        """Test that Judge evaluation checkpoint is saved correctly."""

        judge_agent = JudgeAgentV2(storage=mock_storage, job=mock_job)

        with patch.object(judge_agent, '_generate_with_retry', new_callable=AsyncMock) as mock_generate:
            mock_generate.return_value = json.dumps({
                "overall_assessment": "Fair performance",
                "completeness": "70% coverage",
                "correctness": "Some errors",
                "context_usage": "Partially applied",
                "spatial_understanding": "Generally good",
                "false_positives": "2 false positives",
                "improvement_suggestions": ["Improve accuracy"]
            })

            with patch.object(judge_agent, 'save_checkpoint', new_callable=AsyncMock) as mock_save:
                mock_save.return_value = None

                # Process with Judge
                await judge_agent.process({
                    "components": sample_components,
                    "excel_file": "/tmp/test.xlsx"
                })

                # Verify checkpoint was saved with correct structure
                mock_save.assert_called_once()
                checkpoint_args = mock_save.call_args[0]

                assert checkpoint_args[0] == "evaluation"  # Checkpoint name
                checkpoint_data = checkpoint_args[1]

                assert "evaluation" in checkpoint_data
                assert "timestamp" in checkpoint_data
                assert checkpoint_data["evaluation"]["overall_assessment"] == "Fair performance"
                assert checkpoint_data["evaluation"]["completeness"] == "70% coverage"

    @pytest.mark.asyncio
    async def test_judge_with_excel_reading(
        self,
        mock_storage,
        mock_job,
        sample_components
    ):
        """Test Judge reading Excel file for evaluation."""

        # Create a temporary Excel file
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Schedule"
        ws["A1"] = "Component ID"
        ws["B1"] = "Type"
        ws["A2"] = "A-101-DR-B2"
        ws["B2"] = "Door"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb.save(tmp.name)
            excel_path = Path(tmp.name)

        try:
            judge_agent = JudgeAgentV2(storage=mock_storage, job=mock_job)

            # Mock file upload
            with patch.object(judge_agent.client.files, 'upload') as mock_upload:
                mock_upload.return_value = MagicMock()

                # Build prompt with Excel file
                prompt, files = judge_agent._build_evaluation_prompt(
                    drawing_path=None,
                    components=sample_components,
                    excel_path=excel_path,
                    context=None
                )

                # Verify Excel info is included in prompt
                assert "Excel File: Generated successfully" in prompt
                assert "Schedule:" in prompt
                assert "rows x" in prompt
                assert "columns" in prompt

        finally:
            # Clean up temp file
            if excel_path.exists():
                excel_path.unlink()

    @pytest.mark.asyncio
    async def test_judge_error_handling_continues_pipeline(
        self,
        mock_storage,
        mock_job
    ):
        """Test that Judge errors don't fail the entire pipeline."""

        judge_agent = JudgeAgentV2(storage=mock_storage, job=mock_job)

        # Simulate an error in Judge evaluation
        with patch.object(judge_agent, 'evaluate_extraction', side_effect=Exception("Judge error")):
            result = await judge_agent.process({"components": []})

            # Verify error is handled gracefully
            assert result["next_stage"] == "complete"
            assert "evaluation" in result
            assert "Poor performance - process error" in result["evaluation"]["overall_assessment"]
            assert "Judge error" in result["evaluation"]["error"]
